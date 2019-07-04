from openpyxl import Workbook
from param import Params
paramater = Params().get_main_args()
class data_save_read():
    def write_date(*args):
        outwb = Workbook()
        outws = outwb.create_sheet('Input Data')
        col = 0
        for arg in args:
            col += 1
            for i in range(len(arg)):
                outws.cell(row=i+1, column=col).value = arg[i]
        outwb.save(paramater.save_data_xlsx)
        return None
    def read_data(name):
        import xlrd
        fname = name
        bk = xlrd.open_workbook(fname)
        all_data = []
        try:
            sh = bk.sheet_by_name('Input Data')
            nrows = sh.nrows # hang
            ncols = sh.ncols # lie
            for i in range(0, nrows):
                every_row_data = []
                for j in range(0, ncols):
                    row_data  = sh.cell_value(i, j)
                    every_row_data.append(row_data)
                all_data.append(every_row_data)
        except:
            print('No sheet in {} named Sheets'.format(fname))
        return all_data
if __name__ == '__main__':

    data_save_read.write_date([1, 2, 3], [7, 8, 9])
    data = data_save_read.read_data(paramater.save_data_xlsx)
    for i in range(len(data)):
        print('data', data[i])